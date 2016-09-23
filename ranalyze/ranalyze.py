#!/usr/bin/env python
"""
Analysis tool to traverse a set of subreddits extracting post information including:
 - Post title
 - Post date
 - Post URL (if not a self post)
 - Number of upvotes/downvotes
"""
from argparse import (ArgumentParser, FileType)
import datetime
from itertools import chain
from re import match
import praw
import openpyxl

DATE_REGEX = r"^\d{4}-\d{2}-\d{2}$"
SUBREDDIT_SANITIZE_REGEX = r"^(?:\/?r?\/?)([^\/]*)\/?$"

INVALID_DATE_ERROR = "{} is not a valid date"
INVALID_DATE_FORMAT_ERROR = ("{} is not a valid date format. "
                             "Proper format is YYYY-MM-DD")
MALFORMED_SHEET_ERROR = "Malformed excel sheet"
MISSING_SUBREDDIT_ERROR = ("at least one subreddit specifier "
                           "(-i or -s) is required")

CLI_ARGUMENTS = {
    "date range": {
        ("-a", "--after"): {
            "help": "only analyze posts on or after this date",
            "required": True,
            "type": str
        },
        ("-b", "--before"): {
            "help": ("only analyze posts on or before "
                     "this date, defaults to today"),
            "default": datetime.date.today().isoformat(),
            "type": str
        }
    },
    "subreddit selection": {
        ("-i", "--input-file"): {
            "action": "append",
            "help": ("name of file containing space-delimited "
                     "list of subreddits to analyze"),
            "nargs": "+",
            "type": FileType()
        },
        ("-s", "--subreddit"): {
            "action": "append",
            "help": ("subreddit to analyze, may be a single value, "
                     "or a space-delimited list"),
            "nargs": "+",
            "type": str
        }
    },
    "output": {
        ("-o", "--output-file"): {
            "help": ("analysis data will be written "
                     "to this file (must be *.xlsx)"),
            "required": True,
            "type": FileType(mode='w')
        }
    }
}


class Post:
    """
    Post
    """

    EXCEL_COLUMN_ORDER = ('permalink',
                          'url',
                          'num_comments',
                          'upvotes',
                          'downvotes',
                          'title',
                          'time_submitted',
                          'time_retrieved')

    def __init__(self, permalink, url, num_comments,
                 upvotes, downvotes, title,
                 time_submitted, time_retrieved):
        self.permalink = permalink
        self.url = url
        self.num_comments = num_comments
        self.upvotes = upvotes
        self.downvotes = downvotes
        self.title = title
        self.time_submitted = time_submitted
        self.time_retrieved = time_retrieved

    def as_dict(self):
        return {key: getattr(self, key) for key in Post.EXCEL_COLUMN_ORDER}


def fetch_data(subreddit_set, after, before):
    """
    :param subreddit_set: set of subreddits from which to retrieve data
    :param after: ignores posts before this datetime
    :param before: ignores posts after this datetime
    :return: yields individual reddit posts one at a time
    :rtype: collections.Iterable[Post]
    """
    reddit = praw.Reddit(user_agent="Documenting ecig subs")
    for subreddit_name in subreddit_set:
        subreddit = reddit.get_subreddit(subreddit_name)
        for post_data in subreddit.get_new(limit=None):
            post_date = datetime.datetime.fromtimestamp(post_data.created_utc)
            if post_date.date() > before:
                continue
            if post_date.date() < after:
                continue
            pretty_submitted_date = post_date.isoformat()
            pretty_now_date = datetime.datetime.utcnow().isoformat()
            post = Post(
                permalink=post_data.permalink,
                url=post_data.url,
                num_comments=post_data.num_comments,
                upvotes=post_data.ups,
                downvotes=post_data.downs,
                title=post_data.title,
                time_submitted=pretty_submitted_date,
                time_retrieved=pretty_now_date
            )
            yield(subreddit_name, post)


def parse_args():
    """
    Parse command line arguments. Generates a set of subreddits to analyze, an
    inclusive date range, and an output file.

    All arguments are validated. Errors may be raised for:
    - Invalid Date Format (e.g. 9/4/2016)
    - Invalid Date (e.g. 2/30/2016)
    - Missing -a or --after argument
    - Non-existent input file
    - No specified subreddits
    - No write permission for output file

    Subreddits are sanitized to remove extraneous characters:
    - leading path information (i.e. /, r/, /r/)
    - trailing slashes

    Subreddits are returned as a set of strings
    Date range is returned as a tuple of datetime.date objects
    Output filename is returned as a string

    :return: set of subreddits, date range (after, before), output filename
    :rtype: set, tuple, str
    """

    def validate_date(date_str):
        """
        Validate format of the date string, and convert to datetime.date object

        Raises an error if:
        - date_str is in an invalid format
        - date_str represents an invalid date

        :param date_str: ISO 8601 formatted date, specifically YYYY-MM-DD
        :return: converted datetime.date object
        :rtype: datetime.date
        """
        if match(DATE_REGEX, date_str) is None:
            parser.error(INVALID_DATE_FORMAT_ERROR.format(date_str))
        try:
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            parser.error(INVALID_DATE_ERROR.format(date_str))
        return date

    parser = ArgumentParser()

    # Add the arguments to the parser

    for group_name, args in CLI_ARGUMENTS.items():
        group = parser.add_argument_group(group_name)
        for flags, options in args.items():
            group.add_argument(*flags, **options)

    args = parser.parse_args()

    # Output file was only opened to verify write access

    args.output_file.close()

    # Validate date arguments and convert to datetime.date

    date_range = tuple(validate_date(arg) for arg in (args.after, args.before))

    # At least one of -s or -i is required

    if args.input_file is None and args.subreddit is None:
        parser.error(MISSING_SUBREDDIT_ERROR)

    # Parse subreddit set

    subreddits = set()

    if args.input_file is not None:
        for file in chain(*args.input_file):
            file_contents = file.read()
            if file_contents.endswith("\n"):
                file_contents = file_contents[:-1]
            file_set = file_contents.split(" ")
            subreddits.update(tuple(file_set))

    if args.subreddit is not None:
        args.subreddit = chain(*args.subreddit)  # flatten 2d subreddit list
        subreddits.update(tuple(args.subreddit))

    # Sanitize subreddit names

    sanitized_subreddits = set(match(SUBREDDIT_SANITIZE_REGEX, sub).group(1)
                               for sub in subreddits)

    return sanitized_subreddits, date_range, args.output_file.name


def wexcel(wb, columns, post):
    """
    append post to end of sheet corresponding to subreddit, 
    making a new sheet if no sheet exists
    :param wb: openpyxl workbook to append post to
    :param columns: column order to use in excel sheet
    :param post: post to append in format (subreddit_name, post) where post is a dict
    """
    
    def columnize(n):
        """
        takes and integer n and returns the string corresponding to the nth column in an excel sheet
        :param n: integer
        :return: string corresponding to the nth column in an excel sheet
        :rtype: str
        """
        column = ''
        if n == 0:
            return 'A'
        #mapping off by one, so add one to correct mapping
        n += 1
        while n > 0:
            column = chr(ord('A') + n % 26 - (1 if n < 25 else 0)) + column
            n -= n % 26
            n //= 26
        return column
    #add post to existing sheet
    if post[0] in wb.sheetnames:
        sheet = wb[post[0]]
        x = 0
        row = sheet.max_row + 1
        while x < len(columns):
            cell = '{}{}'.format(columnize(x), row)
            sheet[cell] = post[1][columns[x]]
            x += 1
    else:
        #sheet did not exist, so create one and add column headings and the post
        sheet = wb.create_sheet(post[0])
        x = 0
        while x < len(columns):
            sheet['{}1'.format(columnize(x))] = columns[x]
            sheet['{}2'.format(columnize(x))] = post[1][columns[x]]
            x += 1


def main():
    """

    :return:
    """
    subreddits, date_range, output_file_name = parse_args()

    workbook = openpyxl.Workbook()

    for subreddit, post in fetch_data(subreddits, *date_range):
        wexcel(workbook, Post.EXCEL_COLUMN_ORDER, (subreddit, post.as_dict()))

    workbook.save(output_file_name)

if __name__ == "__main__":
    main()
